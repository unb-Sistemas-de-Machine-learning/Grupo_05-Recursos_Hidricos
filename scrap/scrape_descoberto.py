# ---------------------------------------- BIBLIOTECAS ----------------------------------------
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
import re
import time
import sys

# ------------------------------------------ PARÂMETROS -----------------------------------------
URL = "https://app.powerbi.com/view?r=eyJrIjoiYTQ1YWYyZjQtN2IxNy00NzFhLTg1OTUtNWMxZDU3YmU5MzYxIiwidCI6IjczZGJmMTMyLWE0YTQtNDkwMy1hYzI2LWJiMjhmY2Y3NDdhNCJ9"
TITULO_VISUAL = "HISTÓRICO DESCOBERTO"
OUTPUT_XLSX = "reservatorios_df-adasa.xlsx"   # salva/atualiza aba Descoberto
SHEET_NAME = "Descoberto"

TIMEOUT = 60
HEADLESS = True  # coloque False para ver o navegador

# Rolagem / materialização
SCROLL_PAUSE = 0.18
BIG_JUMP_PAUSE = 0.45
CYCLES_NO_GROWTH_LIMIT = 4
MAX_TOTAL_TIME_ALL = 3600
MAX_TOTAL_TIME_LIMITED = 600
PAGE_DOWN_BURST = 8
SMALL_STEP = 220

# Progresso
PROGRESS_EVERY = 50
HEARTBEAT_EVERY_SEC = 30

# ------------------------------------------ MENU ----------------------------------------------
def prompt_row_limit():
    print("\n=== Coleta de linhas (DESCoberto) ===")
    print("Digite um número (ex.: 300) ou 'todas' para coletar tudo.\n")
    while True:
        val = input("Quantas linhas deseja coletar? [número|todas]: ").strip().lower()
        if val in ("todas", "all", "a", "*"):
            print("→ Coletando TODAS as linhas (pode demorar).")
            return None
        if val.isdigit() and int(val) > 0:
            n = int(val)
            print(f"→ Coletando até {n} linhas.")
            return n
        print("Entrada inválida. Tente novamente…")

# ------------------------------------------ SELENIUM -------------------------------------------
def build_driver():
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1680,2600")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    service = Service()
    return webdriver.Chrome(service=service, options=opts)

def wait_ready(driver):
    WebDriverWait(driver, TIMEOUT).until(EC.presence_of_element_located((By.TAG_NAME, "h3")))
    time.sleep(2)

def back_to_top(driver):
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.3)

def find_visual_by_title(driver, title_text):
    back_to_top(driver)
    h3 = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.XPATH, f"//h3[normalize-space()='{title_text}']"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", h3)
    time.sleep(0.3)
    el = h3
    for _ in range(12):
        if el.find_elements(By.XPATH, ".//div[@role='grid' or @role='table']"):
            return el
        el = el.find_element(By.XPATH, "./..")
    return h3.find_element(By.XPATH, "./ancestor::*[1]")

def get_grid(visual):
    grids = visual.find_elements(By.XPATH, ".//div[@role='grid' or @role='table']")
    return grids[0] if grids else visual

def _is_scrollable(driver, el):
    try:
        sh = driver.execute_script("return arguments[0].scrollHeight;", el)
        ch = driver.execute_script("return arguments[0].clientHeight;", el)
        return isinstance(sh, (int, float)) and isinstance(ch, (int, float)) and sh > ch + 2
    except Exception:
        return False

def _best_descendant_scrollable(driver, root):
    return driver.execute_script("""
        const root = arguments[0];
        let best = null, bestScore = 0;
        const walker = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
        while (walker.nextNode()) {
            const el = walker.currentNode;
            const sh = el.scrollHeight || 0;
            const ch = el.clientHeight || 0;
            if (sh > ch + 2) {
                const score = sh - ch;
                if (score > bestScore) { best = el; bestScore = score; }
            }
        }
        return best;
    """, root)

def get_scroll_host(driver, grid, visual):
    parent = grid
    for _ in range(12):
        try:
            parent = parent.find_element(By.XPATH, "./..")
        except Exception:
            break
        if _is_scrollable(driver, parent):
            return parent
    best = _best_descendant_scrollable(driver, visual)
    if best:
        return best
    candidates_xpath = "|".join([
        ".//div[contains(@class,'scrollHost')]",
        ".//div[contains(@class,'virtualScroll')]",
        ".//div[contains(@class,'scroll-wrapper')]",
        ".//div[contains(@class,'innerContainer')]",
        ".//div[contains(@class,'scroll') and (contains(@style,'overflow') or contains(@class,'v-'))]",
    ])
    for c in visual.find_elements(By.XPATH, candidates_xpath):
        if _is_scrollable(driver, c):
            return c
    return visual

# --------------- Cabeçalhos (com fallback) ----------------
def read_header_map(grid):
    headers = grid.find_elements(By.XPATH, ".//div[@role='columnheader']")
    mapped = []
    for h in headers:
        name = (h.text or "").replace("\u00A0", " ").strip()
        col_attr = h.get_attribute("column-index")
        if not name or "seleção" in name.lower() or "selecionar" in name.lower() or "ícone" in name.lower():
            continue
        if not col_attr:
            continue
        try:
            col_idx = int(col_attr)
        except Exception:
            continue
        mapped.append((name, col_idx))
    mapped.sort(key=lambda t: t[1])

    if not mapped and headers:
        names = []
        for h in headers:
            nm = (h.text or "").replace("\u00A0", " ").strip()
            if nm and "seleção" not in nm.lower() and "selecionar" not in nm.lower():
                names.append(nm)
        mapped = [(n, i) for i, n in enumerate(names)]
    return mapped

# --------------- utilitários de scroll --------------------
def _at_bottom(driver, el):
    try:
        sh = driver.execute_script("return arguments[0].scrollHeight;", el)
        st = driver.execute_script("return arguments[0].scrollTop;", el)
        ch = driver.execute_script("return arguments[0].clientHeight;", el)
        return sh and ch and (st + ch >= sh - 2)
    except Exception:
        return False

def _jump_top(driver, host):
    driver.execute_script("arguments[0].scrollTop = 0;", host)
    time.sleep(BIG_JUMP_PAUSE)

def _jump_bottom(driver, host):
    driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", host)
    time.sleep(BIG_JUMP_PAUSE)

def _page_downs(driver, host, times):
    ActionChains(driver).move_to_element(host).click(host).perform()
    for _ in range(times):
        ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
        time.sleep(SCROLL_PAUSE/2)

def _page_ups(driver, host, times):
    ActionChains(driver).move_to_element(host).click(host).perform()
    for _ in range(times):
        ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
        time.sleep(SCROLL_PAUSE/2)

def _micro_scroll(driver, host, steps=8):
    for _ in range(steps):
        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollTop + arguments[1];",
                              host, SMALL_STEP)
        time.sleep(SCROLL_PAUSE)

# --------------- coleta + PROGRESSO -----------------------
def stream_rows_full(driver, grid, host, header_map, tabela_nome="", row_limit=None, max_total_time=900):
    headers = [n for n, _ in header_map]
    if not headers:
        print(f"[{tabela_nome}] ERRO: não foi possível detectar cabeçalhos.")
        return [], headers
    colidx_to_name = {idx: name for name, idx in header_map} if header_map else {}
    data_col_name = next((n for n in headers if n.lower().startswith("data")), None)

    seen = set()
    out = []

    start_total = time.time()
    cycles_no_growth = 0
    last_progress_print = 0
    last_heartbeat = 0
    last_seen_date = None

    def log_progress(force=False):
        nonlocal last_progress_print, last_heartbeat
        now = time.time()
        if force or len(seen) >= last_progress_print + PROGRESS_EVERY:
            last_progress_print = len(seen)
            meta = f"/{row_limit}" if row_limit else ""
            msg = f"[{tabela_nome}] coletadas {len(seen)}{meta} linhas"
            if last_seen_date: msg += f" | última: {last_seen_date}"
            print(msg); sys.stdout.flush()
        if now - last_heartbeat >= HEARTBEAT_EVERY_SEC:
            last_heartbeat = now
            elapsed = int(now - start_total)
            hb = f"[{tabela_nome}] (heartbeat) {len(seen)} linhas | {elapsed}s decorrido"
            if last_seen_date: hb += f" | última: {last_seen_date}"
            print(hb); sys.stdout.flush()

    def collect_visible():
        nonlocal last_seen_date
        added = 0
        rows = grid.find_elements(By.XPATH, ".//div[@role='row' and .//div[@role='gridcell']]")
        for r in rows:
            cells = r.find_elements(By.XPATH, ".//div[@role='gridcell']")
            if not cells: continue
            row_map = {name: "" for name in headers}
            for c in cells:
                idx_attr = c.get_attribute("column-index")
                txt = c.text.strip().replace("\u00A0", " ")
                if idx_attr and idx_attr.isdigit() and colidx_to_name:
                    name = colidx_to_name.get(int(idx_attr))
                    if name: row_map[name] = txt
            if any(row_map[n] == "" for n in headers):
                ordered = [c.text.strip().replace("\u00A0", " ") for c in cells]
                k = 0
                for name in headers:
                    if row_map[name] == "" and k < len(ordered):
                        row_map[name] = ordered[k]; k += 1
            dv = (row_map.get(data_col_name, "") if data_col_name else "").strip()
            if not dv or "selecionar" in dv.lower(): continue
            if not re.match(r"^\d{2}\.\d{2}\.\d{2}$", dv): continue
            last_seen_date = dv
            if dv not in seen:
                seen.add(dv)
                out.append(row_map)
                added += 1
                if len(out) <= 3:
                    print(f"[{tabela_nome}] exemplo linha {len(out)}: {row_map}")
                if row_limit and len(seen) >= row_limit:
                    break
        if added: log_progress()
        return added

    print(f"[{tabela_nome}] cabeçalhos detectados: {headers}")
    print(f"[{tabela_nome}] iniciando coleta…"); sys.stdout.flush()

    while True:
        if (time.time() - start_total) > max_total_time:
            print(f"[{tabela_nome}] limite de tempo atingido ({max_total_time}s)."); break
        if row_limit and len(out) >= row_limit:
            print(f"[{tabela_nome}] atingiu o limite definido ({row_limit})."); break

        before = len(out)

        _jump_top(driver, host); collect_visible()
        for _ in range(1_000_000):
            if row_limit and len(out) >= row_limit: break
            _page_downs(driver, host, times=PAGE_DOWN_BURST)
            time.sleep(SCROLL_PAUSE); collect_visible()
            _micro_scroll(driver, host, steps=6); collect_visible()
            if _at_bottom(driver, host): break

        _jump_bottom(driver, host); collect_visible()
        for _ in range(1_000_000):
            if row_limit and len(out) >= row_limit: break
            _page_ups(driver, host, times=PAGE_DOWN_BURST)
            time.sleep(SCROLL_PAUSE); collect_visible()
            st = driver.execute_script("return arguments[0].scrollTop;", host)
            if st == 0: break

        after = len(out)
        cycles_no_growth = cycles_no_growth + 1 if after == before else 0
        log_progress(force=True)
        if cycles_no_growth >= CYCLES_NO_GROWTH_LIMIT:
            break

    if any(h.lower().startswith("data") for h in headers):
        def _parse(d):
            try: return pd.to_datetime(d, format="%d.%m.%y")
            except Exception: return pd.NaT
        out.sort(key=lambda r: _parse(r.get("Data", "")), reverse=True)

    print(f"[{tabela_nome}] FINALIZADO com {len(out)} linhas coletadas."); sys.stdout.flush()
    return out, headers

def scrape_table(driver, row_limit):
    visual = find_visual_by_title(driver, TITULO_VISUAL)
    grid = get_grid(visual)
    WebDriverWait(driver, TIMEOUT).until(EC.presence_of_element_located((By.XPATH, ".//div[@role='columnheader']")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", grid)
    time.sleep(0.2)
    host = get_scroll_host(driver, grid, visual)
    header_map = read_header_map(grid)
    rows, headers = stream_rows_full(
        driver, grid, host, header_map,
        tabela_nome=SHEET_NAME,
        row_limit=row_limit,
        max_total_time=(MAX_TOTAL_TIME_ALL if row_limit is None else MAX_TOTAL_TIME_LIMITED),
    )
    return headers, rows

# ---------------------------------------- EXCEL -----------------------------------------------
def write_sheet(ws, title, headers, rows):
    ncol = len(headers)
    if ncol == 0: return
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=title)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="1E88E5")
    c.font = Font(color="FFFFFF", bold=True, size=14)

    head_fill = PatternFill("solid", fgColor="90CAF9")
    head_font = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, row in enumerate(rows, start=3):
        for j, h in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=row.get(h, ""))

    for j, h in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(12, len(h) + 2)

# --------------------------------------- EXECUÇÃO ----------------------------------------------
if __name__ == "__main__":
    row_limit = prompt_row_limit()  # None = todas

    # abre ou cria workbook
    try:
        wb = openpyxl.load_workbook(OUTPUT_XLSX)
    except Exception:
        wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    driver = build_driver()
    driver.get(URL)
    wait_ready(driver)

    headers, rows = scrape_table(driver, row_limit)
    write_sheet(ws, TITULO_VISUAL, headers, rows)

    driver.quit()
    wb.save(OUTPUT_XLSX)
    print(f"\nArquivo Excel salvo: {OUTPUT_XLSX}")
